import re
import time
from openpyxl import load_workbook

#ブックを取得
book = load_workbook(filename='Excelファイルのパス')

#シートを取得
sheet = book['出力したいシート名']

#excelに書き込む関数
def output_excel(recive_data, week_time, gc_list):
    ro = 3
    for i in range(len(gc_list)):
        col = 4
        for j in range(len(week_time)):
            output_data = None
            try:
                #strage_data = _def_list[gc_list[i]][week_time[j]]
                strage_data = recive_data.get(gc_list[i], {}).get(week_time[j]) #学年・コマをkeyとして、リストから授業名を取得する。
                if strage_data != None:
                    output_data = '\n'.join(strage_data) #リストから複数のデータ(授業名)があるとき、改行を使って文字列を連結させる。
                    sheet.cell(row=ro, column=col).value = output_data #excelのro行(row)、col列(column)にoutput_dataを書き込む
            except KeyError:
                pass

            col += 1
        ro += 1
    
    #書き込んだ内容を保存(bookの変更を('xxx/yyy.xlsx')に保存する)
    book.save('Excelファイルのパス') 
            
#output_excel(abc, week_time, gc_list)


#book.close()